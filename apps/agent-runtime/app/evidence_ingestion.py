from __future__ import annotations

import csv
import json
import os
import subprocess
import tempfile
import time
import uuid
from dataclasses import dataclass
from pathlib import Path
from typing import Any
from urllib.error import HTTPError, URLError
from urllib.request import Request, urlopen

from .video_sampling import VideoSamplingError, sample_video_frames, video_duration_ms


class IngestionError(Exception):
    """A safe, retryable ingestion failure."""


@dataclass(frozen=True)
class ClaimedJob:
    job_id: str
    attempts: int
    artifact_id: str
    project_id: str
    organization_id: str
    evidence_type: str
    filename: str
    media_type: str
    storage_key: str
    storage_version: str | None
    checksum_sha256: str


class EvidenceIngestionWorker:
    def __init__(self, environment: dict[str, str] | None = None) -> None:
        self.env = environment or dict(os.environ)
        self.supabase_url = self._required("NEXT_PUBLIC_SUPABASE_URL").rstrip("/")
        self.service_key = self._required("SUPABASE_SERVICE_ROLE_KEY")
        self.transcription_model = self.env.get("EVIDENCE_TRANSCRIPTION_MODEL")
        self.openai_key = self.env.get("OPENAI_API_KEY")
        default_provider = (self.env.get("EVIDENCE_TRANSCRIPTION_PROVIDER") or "openai").lower()
        self.audio_transcription_provider = (
            self.env.get("EVIDENCE_AUDIO_TRANSCRIPTION_PROVIDER") or default_provider
        ).lower()
        self.video_transcription_provider = (
            self.env.get("EVIDENCE_VIDEO_TRANSCRIPTION_PROVIDER") or default_provider
        ).lower()
        self.modal_transcription_url = self.env.get("EVIDENCE_MODAL_TRANSCRIPTION_URL")
        self.modal_proxy_auth_key = self.env.get("EVIDENCE_MODAL_PROXY_AUTH_KEY")
        self.modal_proxy_auth_secret = self.env.get("EVIDENCE_MODAL_PROXY_AUTH_SECRET")
        self.video_coverage_seconds = int(self.env.get("EVIDENCE_VIDEO_COVERAGE_SECONDS", "15"))
        self.video_max_frames = int(self.env.get("EVIDENCE_VIDEO_MAX_FRAMES", "60"))
        self.video_scene_threshold = float(self.env.get("EVIDENCE_VIDEO_SCENE_THRESHOLD", "0.2"))

    def _required(self, name: str) -> str:
        value = self.env.get(name)
        if not value:
            raise IngestionError(f"{name} is not configured for evidence ingestion.")
        return value

    def run_once(self) -> bool:
        self._purge_expired()
        claimed = self._claim()
        if claimed is None:
            return False
        try:
            with tempfile.TemporaryDirectory(prefix="tacit-evidence-") as directory:
                source = Path(directory) / "source"
                self._download(claimed.storage_key, source)
                self._scan(source)
                self._update_artifact(
                    claimed.artifact_id,
                    {"status": "processing", "scan_status": "clean", "failure_reason": None},
                )
                extractions = self._extract(claimed, source, Path(directory))
                self._save_extractions(claimed, extractions)
            self._update_artifact(
                claimed.artifact_id,
                {"status": "ready", "scan_status": "clean", "failure_reason": None},
            )
            self._complete_job(claimed.job_id, "succeeded")
        except Exception as error:  # Worker errors are made retryable and never leak to customers.
            self._retry_or_fail(claimed, error)
        return True

    def _purge_expired(self) -> None:
        now = time.strftime("%Y-%m-%dT%H:%M:%SZ", time.gmtime())
        artifacts = self._request(
            f"/rest/v1/evidence_artifacts?retention_expires_at=lt.{now}&status=neq.deleted&select=id,storage_key"
        )
        for artifact in artifacts:
            self._request(
                "/storage/v1/object/tacit-artifacts",
                method="DELETE",
                body={"prefixes": [artifact["storage_key"]]},
            )
            self._request(
                f"/rest/v1/evidence_artifacts?id=eq.{artifact['id']}",
                method="PATCH",
                body={"status": "deleted", "deleted_at": now, "updated_at": now},
            )
            self._request(
                f"/rest/v1/evidence_ingestion_jobs?artifact_id=eq.{artifact['id']}&status=in.(queued,running)",
                method="PATCH",
                body={"status": "cancelled", "completed_at": now, "updated_at": now},
            )

    def _request(
        self,
        path: str,
        *,
        method: str = "GET",
        body: Any | None = None,
        headers: dict[str, str] | None = None,
    ) -> Any:
        payload = None if body is None else json.dumps(body).encode("utf-8")
        request = Request(f"{self.supabase_url}{path}", data=payload, method=method)
        request.add_header("apikey", self.service_key)
        request.add_header("Authorization", f"Bearer {self.service_key}")
        if payload is not None:
            request.add_header("Content-Type", "application/json")
        for key, value in (headers or {}).items():
            request.add_header(key, value)
        try:
            with urlopen(request, timeout=45) as response:
                return (
                    json.loads(response.read().decode("utf-8"))
                    if response.headers.get("content-type", "").startswith("application/json")
                    else response.read()
                )
        except HTTPError as error:
            raise IngestionError(f"Evidence persistence request failed ({error.code}).") from error

    def _claim(self) -> ClaimedJob | None:
        rows = self._request("/rest/v1/rpc/claim_evidence_ingestion_job", method="POST", body={})
        if not rows:
            return None
        row = rows[0]
        return ClaimedJob(**row)

    def _download(self, storage_key: str, destination: Path) -> None:
        request = Request(
            f"{self.supabase_url}/storage/v1/object/tacit-artifacts/{'/'.join(storage_key.split('/'))}"
        )
        request.add_header("apikey", self.service_key)
        request.add_header("Authorization", f"Bearer {self.service_key}")
        try:
            with urlopen(request, timeout=120) as response, destination.open("wb") as output:
                while chunk := response.read(1024 * 1024):
                    output.write(chunk)
        except HTTPError as error:
            raise IngestionError(
                f"Evidence source could not be downloaded ({error.code})."
            ) from error

    def _scan(self, source: Path) -> None:
        """Treat validated uploads as scan-clean for the hackathon deployment."""
        del source

    def _extract(self, job: ClaimedJob, source: Path, directory: Path) -> list[dict[str, Any]]:
        if job.media_type in {"text/plain", "text/markdown"}:
            return [
                {
                    "kind": "text",
                    "content": source.read_text(encoding="utf-8", errors="replace"),
                    "page_start": None,
                    "page_end": None,
                    "time_start_ms": None,
                    "time_end_ms": None,
                    "confidence": 1.0,
                }
            ]
        if job.media_type in {"text/csv", "application/csv"}:
            return self._extract_csv(source)
        if job.media_type == "application/pdf":
            return self._extract_pdf(source)
        if (
            job.media_type
            == "application/vnd.openxmlformats-officedocument.wordprocessingml.document"
        ):
            return self._extract_docx(source)
        if job.media_type == "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet":
            return self._extract_xlsx(source)
        if job.media_type.startswith("image/"):
            return self._extract_image(source, None, None)
        if job.media_type.startswith("audio/"):
            return self._extract_audio(
                source, 0, self._duration_ms(source), self.audio_transcription_provider
            )
        if job.media_type.startswith("video/"):
            return self._extract_video(source, directory)
        raise IngestionError("No extractor is configured for this validated media type.")

    def _extract_pdf(self, source: Path) -> list[dict[str, Any]]:
        from pypdf import PdfReader

        return [
            {
                "kind": "text",
                "content": page.extract_text().strip(),
                "page_start": index,
                "page_end": index,
                "time_start_ms": None,
                "time_end_ms": None,
                "confidence": 0.96,
            }
            for index, page in enumerate(PdfReader(str(source)).pages, start=1)
            if page.extract_text().strip()
        ]

    def _extract_docx(self, source: Path) -> list[dict[str, Any]]:
        from docx import Document

        content = "\n".join(
            paragraph.text
            for paragraph in Document(str(source)).paragraphs
            if paragraph.text.strip()
        )
        return (
            [
                {
                    "kind": "text",
                    "content": content,
                    "page_start": None,
                    "page_end": None,
                    "time_start_ms": None,
                    "time_end_ms": None,
                    "confidence": 0.94,
                }
            ]
            if content
            else []
        )

    def _extract_csv(self, source: Path) -> list[dict[str, Any]]:
        with source.open(encoding="utf-8-sig", newline="") as file:
            rows = list(csv.reader(file))
        content = "\n".join(
            ", ".join(cell.strip() for cell in row)
            for row in rows
            if any(cell.strip() for cell in row)
        )
        return (
            [
                {
                    "kind": "spreadsheet",
                    "content": content,
                    "page_start": None,
                    "page_end": None,
                    "time_start_ms": None,
                    "time_end_ms": None,
                    "confidence": 1.0,
                }
            ]
            if content
            else []
        )

    def _extract_xlsx(self, source: Path) -> list[dict[str, Any]]:
        from openpyxl import load_workbook

        workbook = load_workbook(source, read_only=True, data_only=True)
        result: list[dict[str, Any]] = []
        for sheet in workbook.worksheets:
            rows = [
                ", ".join("" if cell is None else str(cell) for cell in row)
                for row in sheet.iter_rows(values_only=True)
            ]
            content = f"Sheet: {sheet.title}\n" + "\n".join(row for row in rows if row.strip(", "))
            if content.strip():
                result.append(
                    {
                        "kind": "spreadsheet",
                        "content": content,
                        "page_start": None,
                        "page_end": None,
                        "time_start_ms": None,
                        "time_end_ms": None,
                        "confidence": 1.0,
                    }
                )
        return result

    def _extract_image(
        self, source: Path, start_ms: int | None, end_ms: int | None
    ) -> list[dict[str, Any]]:
        # Image ingestion stores OCR only. Multimodal interpretation is a later worker and
        # creates a visual citation when it needs one for the image pixels.
        content, confidence = self._ocr_image(source)
        if not content:
            return []
        return [
            {
                "kind": "ocr",
                "content": content,
                "page_start": None,
                "page_end": None,
                "time_start_ms": start_ms,
                "time_end_ms": end_ms,
                "confidence": confidence,
            }
        ]

    def _ocr_image(self, source: Path) -> tuple[str, float]:
        """Return OCR text plus mean Tesseract word confidence in [0, 1]."""
        import pytesseract
        from pytesseract import Output

        data = pytesseract.image_to_data(str(source), output_type=Output.DICT)
        lines: dict[tuple[int, int, int], list[str]] = {}
        confidences: list[float] = []
        texts = data.get("text", [])
        confs = data.get("conf", [])
        blocks = data.get("block_num", [])
        pars = data.get("par_num", [])
        line_nums = data.get("line_num", [])
        for index, raw in enumerate(texts):
            word = str(raw or "").strip()
            try:
                score = float(confs[index])
            except (IndexError, TypeError, ValueError):
                continue
            # Tesseract uses -1 for empty / non-word cells.
            if not word or score < 0:
                continue
            confidences.append(min(score, 100.0) / 100.0)
            try:
                key = (int(blocks[index]), int(pars[index]), int(line_nums[index]))
            except (IndexError, TypeError, ValueError):
                key = (0, 0, index)
            lines.setdefault(key, []).append(word)
        if not confidences:
            return "", 0.0
        content = "\n".join(" ".join(parts) for _, parts in sorted(lines.items()))
        confidence = sum(confidences) / len(confidences)
        return content, round(confidence, 4)

    def _extract_audio(
        self, source: Path, start_ms: int, end_ms: int, provider: str
    ) -> list[dict[str, Any]]:
        segments = self._transcribe_segments(source, provider)
        return [
            {
                "kind": "transcript",
                "content": content,
                "page_start": None,
                "page_end": None,
                "time_start_ms": max(start_ms, start_ms + segment_start),
                "time_end_ms": min(end_ms, start_ms + segment_end),
                "confidence": 0.82,
            }
            for content, segment_start, segment_end in segments
            if content and segment_end >= segment_start
        ]

    def _extract_video(self, source: Path, directory: Path) -> list[dict[str, Any]]:
        duration = self._duration_ms(source)
        audio = directory / "audio.mp3"
        self._run_command(
            ["ffmpeg", "-y", "-i", str(source), "-vn", "-ac", "1", "-ar", "16000", str(audio)]
        )
        frames = sample_video_frames(
            source,
            directory / "frames",
            duration_ms=duration,
            coverage_seconds=self.video_coverage_seconds,
            max_frames=self.video_max_frames,
            scene_threshold=self.video_scene_threshold,
        )
        result = self._extract_audio(audio, 0, duration, self.video_transcription_provider)
        for frame in frames:
            end_ms = min(frame.time_ms + self.video_coverage_seconds * 1000, duration)
            result.append(
                {
                    "kind": "frame",
                    "content": f"Video frame sampled at {frame.time_ms / 1000:.3f}s for visual analysis.",
                    "page_start": None,
                    "page_end": None,
                    "time_start_ms": frame.time_ms,
                    "time_end_ms": end_ms,
                    "confidence": 1.0,
                }
            )
            result.extend(self._extract_image(frame.path, frame.time_ms, end_ms))
        return result

    def _duration_ms(self, source: Path) -> int:
        try:
            return video_duration_ms(source)
        except VideoSamplingError as error:
            raise IngestionError("Media processing failed safely.") from error

    def _run_command(self, command: list[str]) -> None:
        try:
            subprocess.run(command, check=True, capture_output=True, timeout=120)
        except (OSError, subprocess.CalledProcessError, subprocess.TimeoutExpired) as error:
            raise IngestionError("Media processing failed safely.") from error

    def _transcribe(self, source: Path, provider: str) -> str:
        if provider == "openai":
            return self._transcribe_openai(source)
        if provider == "modal":
            return self._transcribe_modal(source)
        raise IngestionError(
            "Evidence transcription provider must be configured as 'openai' or 'modal'."
        )

    def _transcribe_segments(self, source: Path, provider: str) -> list[tuple[str, int, int]]:
        if provider != "openai":
            transcript = self._transcribe(source, provider)
            return [(transcript, 0, self._duration_ms(source))] if transcript else []
        if not self.openai_key or not self.transcription_model:
            raise IngestionError("Transcription is not configured for evidence ingestion.")
        boundary = f"----Tacit{uuid.uuid4().hex}"
        payload: list[bytes] = []

        def field(name: str, value: str) -> None:
            payload.extend([
                f"--{boundary}\r\n".encode(),
                f'Content-Disposition: form-data; name="{name}"\r\n\r\n'.encode(),
                value.encode(),
                b"\r\n",
            ])

        field("model", self.transcription_model)
        field("response_format", "verbose_json")
        payload.extend([
            f"--{boundary}\r\n".encode(),
            'Content-Disposition: form-data; name="file"; filename="audio.mp3"\r\nContent-Type: audio/mpeg\r\n\r\n'.encode(),
            source.read_bytes(),
            b"\r\n",
            f"--{boundary}--\r\n".encode(),
        ])
        request = Request("https://api.openai.com/v1/audio/transcriptions", data=b"".join(payload), method="POST")
        request.add_header("Authorization", f"Bearer {self.openai_key}")
        request.add_header("Content-Type", f"multipart/form-data; boundary={boundary}")
        try:
            with urlopen(request, timeout=120) as response:
                body = json.loads(response.read().decode())
        except (HTTPError, URLError, json.JSONDecodeError) as error:
            raise IngestionError("OpenAI transcription request failed safely.") from error
        segments = body.get("segments")
        if isinstance(segments, list):
            parsed = [
                (str(segment.get("text", "")).strip(), int(float(segment.get("start", 0)) * 1000), int(float(segment.get("end", 0)) * 1000))
                for segment in segments
                if isinstance(segment, dict) and str(segment.get("text", "")).strip()
            ]
            if parsed:
                return parsed
        transcript = str(body.get("text", "")).strip()
        return [(transcript, 0, self._duration_ms(source))] if transcript else []

    def _transcribe_openai(self, source: Path) -> str:
        if not self.openai_key or not self.transcription_model:
            raise IngestionError("Transcription is not configured for evidence ingestion.")
        boundary = f"----Tacit{uuid.uuid4().hex}"
        payload: list[bytes] = []

        def field(name: str, value: str) -> None:
            payload.extend(
                [
                    f"--{boundary}\r\n".encode(),
                    f'Content-Disposition: form-data; name="{name}"\r\n\r\n'.encode(),
                    value.encode(),
                    b"\r\n",
                ]
            )

        field("model", self.transcription_model)
        field("response_format", "json")
        payload.extend(
            [
                f"--{boundary}\r\n".encode(),
                'Content-Disposition: form-data; name="file"; filename="audio.mp3"\r\nContent-Type: audio/mpeg\r\n\r\n'.encode(),
                source.read_bytes(),
                b"\r\n",
                f"--{boundary}--\r\n".encode(),
            ]
        )
        request = Request(
            "https://api.openai.com/v1/audio/transcriptions", data=b"".join(payload), method="POST"
        )
        request.add_header("Authorization", f"Bearer {self.openai_key}")
        request.add_header("Content-Type", f"multipart/form-data; boundary={boundary}")
        try:
            with urlopen(request, timeout=120) as response:
                return str(json.loads(response.read().decode("utf-8")).get("text", "")).strip()
        except (HTTPError, URLError) as error:
            raise IngestionError("OpenAI transcription request failed safely.") from error

    def _transcribe_modal(self, source: Path) -> str:
        if not (
            self.modal_transcription_url
            and self.modal_proxy_auth_key
            and self.modal_proxy_auth_secret
        ):
            raise IngestionError("Modal transcription is not configured for evidence ingestion.")
        request = Request(
            self.modal_transcription_url.rstrip("/") + "/transcribe",
            data=source.read_bytes(),
            method="POST",
        )
        request.add_header("Content-Type", "application/octet-stream")
        request.add_header("Modal-Key", self.modal_proxy_auth_key)
        request.add_header("Modal-Secret", self.modal_proxy_auth_secret)
        try:
            with urlopen(request, timeout=150) as response:
                body = json.loads(response.read().decode("utf-8"))
                return str(body.get("text", "")).strip()
        except (HTTPError, URLError, json.JSONDecodeError) as error:
            raise IngestionError("Modal transcription request failed safely.") from error

    def _update_artifact(self, artifact_id: str, value: dict[str, Any]) -> None:
        value["updated_at"] = time.strftime("%Y-%m-%dT%H:%M:%SZ", time.gmtime())
        self._request(
            f"/rest/v1/evidence_artifacts?id=eq.{artifact_id}", method="PATCH", body=value
        )

    def _save_extractions(self, job: ClaimedJob, values: list[dict[str, Any]]) -> None:
        if values:
            source_version = job.storage_version or job.checksum_sha256
            self._request(
                "/rest/v1/evidence_extractions",
                method="POST",
                headers={"Prefer": "return=minimal"},
                body=[
                    {
                        **value,
                        "artifact_id": job.artifact_id,
                        "source_artifact_version": source_version,
                    }
                    for value in values
                ],
            )

    def _complete_job(
        self,
        job_id: str,
        status: str,
        error_code: str | None = None,
        error_message: str | None = None,
    ) -> None:
        self._request(
            f"/rest/v1/evidence_ingestion_jobs?id=eq.{job_id}",
            method="PATCH",
            body={
                "status": status,
                "completed_at": time.strftime("%Y-%m-%dT%H:%M:%SZ", time.gmtime()),
                "error_code": error_code,
                "error_message": error_message,
                "updated_at": time.strftime("%Y-%m-%dT%H:%M:%SZ", time.gmtime()),
            },
        )

    def _retry_or_fail(self, job: ClaimedJob, error: Exception) -> None:
        reason = str(error)[:500] or "Evidence ingestion failed."
        if job.attempts >= 5:
            self._update_artifact(
                job.artifact_id,
                {"status": "failed", "scan_status": "failed", "failure_reason": reason},
            )
            self._complete_job(job.job_id, "failed", "ingestion_failed", reason)
            return
        delay_seconds = min(3600, 30 * (2 ** (job.attempts - 1)))
        available_at = time.strftime("%Y-%m-%dT%H:%M:%SZ", time.gmtime(time.time() + delay_seconds))
        self._update_artifact(
            job.artifact_id, {"status": "queued", "scan_status": "pending", "failure_reason": None}
        )
        self._request(
            f"/rest/v1/evidence_ingestion_jobs?id=eq.{job.job_id}",
            method="PATCH",
            body={
                "status": "queued",
                "available_at": available_at,
                "started_at": None,
                "error_code": None,
                "error_message": None,
                "updated_at": time.strftime("%Y-%m-%dT%H:%M:%SZ", time.gmtime()),
            },
        )
